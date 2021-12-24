import dash
import dash_bootstrap_components as dbc
import dash_daq as daq
import numpy as np
import pandas as pd
import plotly.express as px
from dash import dcc
from dash import html
from dash.dependencies import Input, Output
import openpyxl
from gevent import pywsgi

#Covea Colors
bleu='rgb(28,36,79)'
vert='rgb(89,127,109)'
marron='rgb(143,115,50)'
jaune='rgb(255,218,0)'
saumon='rgb(250,147,112)'
rouge='rgb(184,19,30)'
rouge2='rgb(159,25,87)'
orange='rgb(240,154,0)'
gris='rgb(91,105,113)'
turquoise = 'rgba(0,151,153,1)'
tur='#008D80'


c5 = px.colors.qualitative.Pastel
c6 = px.colors.qualitative.T10


c1 = 'rgba(0,134,149,1)'
c2 = 'rgba(204,102,119,1)'
c3 = 'rgba(217,175,107,1)'
c1b = 'rgb(95,158,160)'
c1c = 'rgb(173,216,230)'


# Iris bar figure
color4 = [c1,c2,c3]
#color4 = [c4,c5[1],c5[2]]


# Pie chart
def drawPie():
    return  html.Div([
        dbc.Card(
            dbc.CardBody([
                html.H5("Repartition de la collecte par typologie patrimoniale", className = "card-title",style = {'textAlign':'center'}),
                dcc.Graph(id = 'pie_chart', figure = {},
                    config={
                        'displayModeBar': False
                    }
                ) 
            ])
        ,color="dark",inverse=True,style={'border':'none'} #- A ajouter pour colorer une des cartes
        )
    ])



# Radio items du choix de la marque
def drawRadio():
    return  dbc.RadioItems(
            id="radio_value",
            inline=False,
            inputClassName="btn-check",
            inputStyle = {},
            labelClassName="btn btn-BC w-100",
            labelStyle={'color':'white'},
            labelCheckedClassName="active",
            options=[
                {"label": "DAV", "value": "DAV"},
                {"label": "GMF", "value": "GMF"},
                {"label": "MAAF", "value": "MAAF"},
                {"label": "MMA", "value": "MMA"},
                {"label": "MMA Agents", "value": "MMA Agents"},
                {"label": "MMA CAP", "value": "MMA CAP"},
                {"label": "MMA Courtage", "value": "MMA Courtage"}
                ],
            value="DAV",
            style={'padding-right':'1.75vw'}
            )


# Texts
def drawText2():
    return html.Div([
        dbc.Card(
            dbc.CardBody([
                html.Div([
                    html.Div(id = "output_uc", children=[],style={'color':'white','fontSize':"25px"})
                ], style={'textAlign': 'center','fontSize':"10px",'fontWeight': 'bold','color':'white'})
            ]),color='rgb(205,70,70)',inverse=True
        )
    ])

def drawText4():
    return html.Div([
        dbc.Card(
            dbc.CardBody([
                html.Div([
                    html.H3("PRIME MOY"),
                    html.Div(id = "output_prime", children=[],style={'color':'white','fontSize':"25px",'bgcolor':'white'})
                ], style={'textAlign': 'center','fontSize':"10px",'fontWeight': 'bold',
                'color':'white'})
            ])
            ,color=px.colors.qualitative.T10[0],inverse=True,
            #style={'height':'150px'}
        )
    ])

def drawGrad0():
    return daq.GraduatedBar(id='gb0',
    step=3,
    size =180,
    max=65,
    value=62,
    color=turquoise,
    showCurrentValue=False,
    className='progressbar_class',
    style = {"height" : "4vh"}
    )



def drawVal2():
    return html.Div([
        dbc.Card([
            dbc.CardHeader(
                html.Div([
                    html.I(className="fas fa-gem",style={'margin-top':'0.5vw','margin-left':'0.5vw','textAlign':'center','fontSize':'2.5vw','color':'rgba(255,255,255,0.9)','height':'3vw','display':'inline-block'}),
                    html.H3("NBV en % CA", className = "card-title mx-auto",style={'margin-top':'0.5vw','textAlign': 'center','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','display':'inline-block'})
                ],className='d-flex align-items-center justify-content-center'),
            style={'background': 'linear-gradient(to right, rgba(70,130,180,0.0), rgba(70,130,180,0.6))','height':'4.5vw'} 
            ),

            dbc.CardBody([
                html.Div([
                    html.Div([
                        html.P("",style = {'height':'0.00vw'}),
                        html.Div([
                            html.H3("Valeur",style={'fontSize':'2vw','color':'rgba(255,255,255,0.9)','height':'2vw'}),
                        ],className='',style={}),

                        html.Div([
                            html.Div(
                                id = "nbglo",
                                children=[],
                                style={'fontSize':"3vw",'margin-left':'0.0vw','margin-bottom':'0vw','fontWeight':'bold',})
                            ],className='',style={}),

                        html.Div([
                            html.Div(
                                id = "nbeu",
                                children=[],
                                style={'fontSize':"1.25vw",'margin-left':'0vw','margin-bottom':'0vw','fontStyle':'italic'})
                            ],className='',style={}),

                        html.Div([
                            html.Div(
                                id = "nbuc",
                                children=[],
                                style={'fontSize':"1.25vw",'margin-left':'0vw','margin-bottom':'0vw','fontStyle':'italic'})
                            ],className='',style={}),
                        html.Br(),

                    ] ,className='text-center',style={'display':'block','height':'13.5vw','width':'13.5vw','border-radius':'100%',"border":"6px rgba(255,255,255,1) solid",
                    "borderColor":'rgb(59,93,120)'}),
                    
                    html.Div([
                        html.Div([
                            html.Div(id = 'nbpc', children=[],style={'color':color4[0],'fontSize':"1.25vw",'margin-left':'2vw','fontWeight':'bold'}),
                            ],style={}),
                        html.Br(),
                        html.Div([
                            html.Div(id = 'nbgp', children=[],style={'color':color4[1],'fontSize':"1.25vw",'margin-left':'2vw','fontWeight':'bold'}),
                            ],style={}),
                        html.Br(),
                        html.Div([
                            html.Div(id = 'nbpp', children=[],style={'color':color4[2],'fontSize':"1.25vw",'margin-left':'2vw','fontWeight':'bold'}),
                            ],style={}),
                    ])
                ],className='d-flex align-items-center justify-content-center',style={'margin-left':'-3vw','display':'inline-block'})
            ],className='card-body d-flex align-items-center justify-content-center',style={'display':'inline-block'})
            
        ],color='primary',inverse=True,style={'border':'none','height':'20vw'})
    ])



tooltip_age = dbc.Tooltip("Age moyen des verseurs",target="tooltip_age",placement='left')
tooltip_turnover = dbc.Tooltip("Taux de sortie annuel moyen (rachats + décès)",target="tooltip_turnover",placement='left')
tooltip_activ = dbc.Tooltip("Part de contrats ayant effectué au moins un versement dans l'année",target="tooltip_activ",placement='left')
tooltip_prime = dbc.Tooltip("Prime moyenne annuelle (hors VP)",target="tooltip_prime",placement='bottom')


colorx='linear-gradient(to right, rgba(66,28,82,0.0), rgba(66,28,82,0.6))'

# Data
dfcarac = pd.read_csv('df_carac.csv',sep=';',decimal=',')
dfmp = pd.read_csv('df_mp.csv',sep=';',decimal=',')
dfmpmk = pd.read_csv('df_mp_marque.csv',sep=';',decimal=',')
dfmaps = pd.read_csv('df_maps.csv',sep=';',decimal=',')
dftable = pd.read_csv('df_table.csv',sep=';',decimal=',',encoding='latin1')
dfcompl = openpyxl.load_workbook('df.xlsx')
dfpvfp = openpyxl.load_workbook('df_pvfp.xlsx')
dfcoll=openpyxl.load_workbook('df_coll.xlsx')

# Build App
FA = "https://use.fontawesome.com/releases/v5.15.1/css/all.css"
app = dash.Dash(external_stylesheets=[dbc.themes.FLATLY,FA],
meta_tags=[
        {"name": "viewport", "content": "width=device-width, initial-scale=1"},
    ],)
server=app.server

app.layout = html.Div([
    dbc.Card(
        dbc.CardBody([
            dbc.Row([

# Barre de navigation
                dbc.Col([
                    
                    dbc.Card([
                            html.Div([
                                html.Img(src = app.get_asset_url('logo.png'),style = {'height':'7vw','margin-top':'1vw','margin-bottom':'1vw'}),
                                html.Br(),
                                html.H3("Direction Assurance Vie", style = {'fontSize':'2vw','textAlign':'center'}),
                                html.Br(),
                                dbc.CardBody([
                                html.H3("Key Performance Indicators",className='text-center',style = {'fontSize':"1.25vw",'color':marron,'fontWeight':'bold'}),
                                ]),
                                html.Img(src = app.get_asset_url('computer.png'),style = {'height':'7vw'}),
                                
                                html.Br(),
                                html.P("Choisir un périmètre", style = {'fontSize':"1.5vw"}),

                                
                                drawRadio(),
                                

                            ],style={'textAlign':'center','color':'White'}),
                    ],className= 'align-items-center',color="primary",inverse=True,style={'height':'65.6vw','border':'none'})
                ], width=2),

# Contenu                
                dbc.Col([

# 2 premieres liones
                    dbc.Row([
                        dbc.Col([
                            
                            dbc.Row([
                                dbc.Col([
                                    dbc.Row([
                                        html.Div([
                                            dbc.Card([
                                                dbc.CardHeader(
                                                    html.Div([
                                                        html.I(className="fas fa-coins",style={'margin-top':'0.5vw','margin-left':'0.5vw','textAlign':'center','fontSize':'2.5vw','color':'rgba(255,255,255,0.9)','height':'3vw','display':'inline-block'}),
                                                        html.H3("Chiffre d'Affaire", className = "card-title mx-auto",style={'margin-top':'0.5vw','textAlign': 'center','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','display':'inline-block'}),
                                                        ],className='d-flex align-items-center justify-content-center'),
                                                    style={'background': 'linear-gradient(to right, rgba(0,151,153,0.0), rgba(0,151,153,0.6))','height':'4.5vw'} 
                                                    ),
                                                dbc.CardBody([
                                                    html.Div([
                                                        html.Div(
                                                            id = "output_container",
                                                            children=[],
                                                            style={'fontSize':"2.5vw"})
                                                        ],style={'textAlign': 'center','fontWeight':'bold'}),
                                                    html.Br(),
                                                    
                                                    
                                                    
                                                    html.Div("Taux UC",className='d-flex justify-content-center',style = {'fontSize':"2vw",'textAlign':'center'}),
                                                    html.H3("",style={'fontSize':'0.5vw'}),
                                                    html.Div(drawGrad0(),className='d-flex justify-content-center',style={'width':'120%','align':'center'}),
                                                    html.Div(id = "output_uc", children=[],className='d-flex justify-content-center',style={'textAlign': 'center','fontSize':"2.0vw",'fontWeight':'bold'})
                                                    
                                                    ],className='card-body justify-content-center align-items-center')],color="primary",inverse=True,style={'height':'20vw','border':'none'} #- A ajouter pour colorer une des cartes
                                            )
                                        ])
                                    ]),
                                    html.Br(),
                                    dbc.Row([drawVal2(),])
                                ],width=4),

                                dbc.Col([
                                    dbc.Row([
                                        html.Div([
                                            dbc.Card([
                                                dbc.CardHeader(
                                                    html.Div([
                                                        
                                                        html.I(className="fas fa-chart-pie",style={'margin-top':'0.5vw','margin-left':'0.5vw','textAlign':'center','fontSize':'2.5vw','color':'rgba(255,255,255,0.9)','height':'3vw','display':'inline-block'}),
                                                        html.H3("Répartition par typologie patrimoniale", className = "card-title mx-auto",style={'margin-top':'0.5vw','textAlign':'center','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','display':'inline-block'})
                                                        ],className='d-flex align-items-center justify-content-center'),
                                                    style={'background': 'linear-gradient(to right, rgba(0,151,153,0.0), rgba(0,151,153,0.6))','height':'4.5vw'} 
                                                    ),
                                                dbc.CardBody([
                                                    html.Div([dcc.Graph(id = 'pie_chart', figure = {},config={'displayModeBar': False},
                                                    style={'height':'18vw'})],
                                                    
                                                    ),                               
                                                    ])
                                                ],color="primary",inverse=True,style={'border':'none','height':'20vw'} #- A ajouter pour colorer une des cartes
                                            )
                                        ])
                                    ]),
                                    html.Br(),
                                    dbc.Row([
                                        html.Div([
                                            dbc.Card([
                                                dbc.CardHeader(
                                                    html.Div([
                                                        html.Img(src = app.get_asset_url('famille2.png'),style = {'margin-top':'0.25vw','margin-bottom':'0.25vw','margin-left':'0.5vw','textAlign':'center','fontSize':'2.5vw','color':'rgba(255,255,255,0.9)','height':'3vw','display':'inline-block'}),
                                                        html.H3("NBV - Déclinaison par âge", className = "card-title mx-auto",style={'margin-top':'0.5vw','textAlign':'center','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','display':'inline-block'})
                                                        ],className='d-flex align-items-center justify-content-center'),
                                                    style={'background': 'linear-gradient(to right, rgba(70,130,180,0.0), rgba(70,130,180,0.6))','height':'10vw'} 
                                                    ),
                                                dbc.CardBody([
                                                    
                                                    dcc.Graph(id = 'scatter_chart', figure = {},config={'displayModeBar': False},
                                                    style={'margin-left':'1vw','height':'20vw'})
                                                    
                                                    ])
                                            ],color='primary',inverse=True,style={'border':'none','height':'20vw'}
                                            )
                                        ])
                                    ]),
                                    html.Br(),
                                ],width=5),

                                dbc.Col([
                                    html.Div([
                                        dbc.Card([
                                            dbc.CardHeader(
                                                html.Div([
                                                    html.I(className="fas fa-user",style={'margin-top':'0.5vw','margin-left':'0.5vw','textAlign':'center','fontSize':'2.5vw','color':'rgba(255,255,255,0.9)','height':'3vw','display':'inline-block'}),
                                                    html.H3("Caractéristiques client", className = "card-title mx-auto",style={'margin-top':'0.5vw','textAlign':'center','fontSize':'1.25vw','color':'rgba(255,255,255,0.9)','display':'inline-block'})
                                                    ],className='d-flex align-items-center justify-content-center'),
                                                style={'background': 'linear-gradient(to right, rgba(220,20,60,0.0), rgba(220,20,60,0.6))','height':'4.5vw'} 
                                                ),
                                            dbc.CardBody([
                                                
                                                    
                                                    html.Div(id='tooltip_age',children=[
                                                        html.Div([
                                                            html.I(className="fas fa-birthday-cake",style={'fontSize':'2.5vw','color':'rgba(255,255,255,0.9)','height':'3vw'}),
                                                            html.H3("Age moyen", style={'fontSize':'1.25vw','color':'rgba(255,255,255,0.9)'}),
                                                        ],style={'textAlign':'center','margin-left':'-1vw','display':'block','width':'50%'}),
                                                        html.Div(id = "output_age", children=[],className='card-body d-flex align-items-center justify-content-end'
                                                        ,style={'margin-left':'0vw','color':'white','fontSize':"1.5vw",'fontWeight':'bold'}),
                                                        tooltip_age
                                                        ],className='card-text d-flex align-items-center justify-content-start',style={'display':'inline-block'}),
                                                    html.P("",style = {'height':'2vw'}),
                                                    html.Div(id='tooltip_turnover',children=[
                                                        html.Div([
                                                            html.I(className="fas fa-sync-alt",style={'fontSize':'2.5vw','color':'rgba(255,255,255,0.9)','height':'3vw'}),
                                                            html.H3("Taux de sortie", style={'fontSize':'1.25vw','color':'rgba(255,255,255,0.9)'})
                                                        ],style={'textAlign':'center','margin-left':'-1vw','display':'block','width':'50%'}),
                                                        html.Div(id = "output_turnover", children=[],className='card-body d-flex align-items-center justify-content-end'
                                                        ,style={'margin-left':'0vw','color':'white','fontSize':"1.5vw",'fontWeight':'bold'}),
                                                        tooltip_turnover,
                                                        ],className='card-text d-flex align-items-center justify-content-start',style={'display':'inline-block'}),
                                                    html.P("",style = {'height':'2vw'}),

                                                    html.Div([
                                                        html.Div(id='tooltip_activ',children=[
                                                            html.I(className="fas fa-credit-card",style={'fontSize':'2.5vw','color':'rgba(255,255,255,0.9)','height':'3vw'}),
                                                            html.H3("Contrats alimentés",style={'fontSize':'1.25vw','color':'rgba(255,255,255,0.9)'})
                                                        ],style={'textAlign':'center','margin-left':'-1vw','display':'block','width':'50%'}),
                                                        html.Div(id = "output_activ", children=[],className='card-body d-flex align-items-center justify-content-end'
                                                        ,style={'margin-left':'-1vw','color':'white','fontSize':"1.5vw",'fontWeight':'bold'}),
                                                        tooltip_activ,
                                                        ],className='card-text d-flex align-items-center justify-content-start',style={'display':'inline-block'}),
                                                    html.P("",style = {'height':'2vw'}),

                                                    html.Div([
                                                        html.Div(id='tooltip_prime',children=[
                                                            html.I(className="fas fa-piggy-bank",style={'fontSize':'2.5vw','color':'rgba(255,255,255,0.9)','height':'3vw'}),
                                                            html.H3("Prime Moyenne", style={'fontSize':'1.25vw','color':'rgba(255,255,255,0.9)'})
                                                        ],style={'textAlign':'center','margin-left':'-1vw','display':'block','width':'50%'}),
                                                        html.Div(id = "output_prime", children=[],className='card-body d-flex align-items-center justify-content-end'
                                                        ,style={'margin-left':'0vw','color':'white','fontSize':"1.5vw",'fontWeight':'bold'}),
                                                        tooltip_prime,
                                                        ],className='card-text d-flex align-items-center justify-content-start',style={'display':'inline-block'}),
                                                    html.Br(),
                                                                                              
                                                ],className='card-body mx-auto',style={'margin-top':'2vw'})],color="primary",inverse=True,style={'border':'none','height':'41.3vw'} #
                                        )
                                    ])
                                ],width=3),
                            ], className='d-flex align-items-stretch justify-content-center'),
                                
                        
                    ]),

# 3e ligne
                    dbc.Container([
                    dbc.Row([
                        dbc.Col([
                            html.Div([
                                dbc.Card([
                                    dbc.CardHeader(
                                        html.Div([
                                            html.I(className="fas fa-user",style={'margin-top':'0.5vw','margin-left':'0.5vw','textAlign':'center','fontSize':'2.5vw','color':'rgba(255,255,255,0.9)','height':'3vw','display':'inline-block'}),
                                            html.H3("Patrimoniaux - 40-50ans", className = "card-title mx-auto",style={'margin-top':'0.5vw','textAlign':'center','fontSize':'1.25vw','color':'rgba(255,255,255,0.9)','display':'inline-block'}),
                                            
                                            ],className='d-flex align-items-center justify-content-center'),
                                        style={'background': 'linear-gradient(to right, rgba(219,112,147,0.0), rgba(219,112,147,0.6))','height':'4.5vw'} 
                                        ),
                                    dbc.CardBody([
                                        html.Div([
                                                html.Div([
                                                    html.P("Valeur",style={'textAlign': 'center','margin-left':'0.2vw','fontSize':'1.2vw','color':'rgba(255,255,255,0.9)','height':'10%'}),
                                                    html.P(id='output_nbvage40',style={'textAlign': 'center','margin-left':'0.2vw','display':'block','fontWeight': 'bold','fontSize':'2vw','color':'rgba(255,255,255,0.9)','height':'20%'}),
                                                    ],className='card-body align-items-center justify-content-right',style={'margin-left':'11vw','margin-top':'2vw','display':'block',"border-radius":"100%","border":"1px rgba(255,255,255) solid",'fontWeight': 'bold','height':'90%','background-color':'rgba(0,0,0,0)','width':'35%'}),

                                                html.Div([
                                                    html.P(id='output_ucage40',style={'textAlign': 'center','margin-left':'0vw','margin-top':'-6vw','display':'block','fontWeight': 'bold','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'20%','width':'45%'}),
                                                    html.P("d'UC",style={'textAlign': 'center','margin-left':'-1.6vw','fontSize':'1vw','margin-top':'-5.4vw','color':'rgba(255,255,255,0.9)','height':'8%'}),
                                                    ],className='card-body d-flex justify-content-left',style={'display':'inline-block','height':'0%'}),

                                                html.Div([
                                                    html.P("Durée :",style={'textAlign': 'center','margin-left':'-0.2vw','fontSize':'1vw','margin-top':'-4.3vw','color':'rgba(255,255,255,0.9)','height':'8%'}),
                                                    html.P(id='output_durage40',style={'textAlign': 'center','margin-left':'-1vw','margin-top':'-4.8vw','display':'block','fontWeight': 'bold','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'20%','width':'45%'}),
                                                    ],className='card-body d-flex justify-content-left',style={'display':'inline-block','height':'16%'}),

                                                html.Div([
                                                    html.I(className="fas fa-chevron-right",style={'textAlign':'center','margin-top':'-7.5vw','margin-left':'8.5vw','fontSize':'2vw','color':'rgba(255,255,255,0.9)','height':'3vw'}),
                                                    ],className='card-body d-flex justify-content-left',style={'display':'inline-block','height':'13%'}),

                                             ],style={'background': 'linear-gradient(to left, rgba(219,112,147,0.0), rgba(20,27,34,0.6))','height':'6vw','margin-top':'-2vw'}),
                                        html.Div([
                                            html.I(className="fas fa-money-bill",style={'textAlign':'center','margin-top':'1.3vw','fontSize':'1.3vw','color':'rgba(255,255,255,0.9)','height':'3vw'}),
                                            html.P(id='output_patage40',style={'textAlign': 'center','margin-left':'-0.5vw','margin-top':'0.6vw','display':'block','fontWeight': 'bold','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'20%','width':'45%'}),
                                            html.P("du CA Confirmé",style={'textAlign': 'center','margin-left':'-1.8vw','fontSize':'1vw','margin-top':'1.10vw','color':'rgba(255,255,255,0.9)','height':'8%'}),
                                            ],className='card-body d-flex justify-content-left',style={'display':'inline-block','height':'13%'}),

                                        html.Div([
                                            html.I(className="fas fa-gas-pump",style={'textAlign':'center','margin-top':'1.8vw','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'3vw'}),
                                            html.P(id='output_nbage40',style={'textAlign': 'center','margin-left':'0.8vw','margin-top':'1.4vw','display':'block','fontWeight': 'bold','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'20%','width':'45%'}),
                                            html.P("contrats",style={'textAlign': 'center','margin-left':'-0.9vw','fontSize':'1vw','margin-top':'1.8vw','color':'rgba(255,255,255,0.9)','height':'8%'}),
                                            ],className='card-body d-flex justify-content-left',style={'display':'inline-block','height':'15%'}),
                                        
                                        html.Div([
                                            html.I(className="fas fa-credit-card mx-auto",style={'textAlign':'center','margin-top':'1.8vw','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'3vw'}),
                                            html.P(id='output_activiteage40',style={'textAlign': 'center','margin-left':'-0.3vw','margin-top':'1.5vw','display':'block','fontWeight': 'bold','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'20%','width':'45%'}),
                                            html.H3("alimenté", className = "card-body d-flex justify-content-center",style={'textAlign': 'center','margin-top':'1.4vw','margin-left':'-7vw','fontSize':'1vw','color':'rgba(255,255,255,0.9)'})
                                            ],className='card-body d-flex justify-content-center',style={'display':'inline-block','height':'15%'}),
                                        ])],color="primary",inverse=True,style={'border':'none','height':'23vw'} #
                                )
                            ])
                        ]),

                        dbc.Col([
                            html.Div([
                                dbc.Card([
                                    dbc.CardHeader(
                                        html.Div([
                                            html.I(className="fas fa-user",style={'margin-top':'0.5vw','margin-left':'0.5vw','textAlign':'center','fontSize':'2.5vw','color':'rgba(255,255,255,0.9)','height':'3vw','display':'inline-block'}),
                                            html.H3("Patrimoniaux - 70-75ans", className = "card-title mx-auto",style={'margin-top':'0.5vw','textAlign':'center','fontSize':'1.25vw','color':'rgba(255,255,255,0.9)','display':'inline-block'}),
                                            
                                            ],className='d-flex align-items-center justify-content-center'),
                                        style={'background': 'linear-gradient(to right, rgba(219,112,147,0.0), rgba(219,112,147,0.6))','height':'4.5vw'} 
                                        ),
                                    dbc.CardBody([
                                        html.Div([
                                                html.Div([
                                                    html.P("Valeur",style={'textAlign': 'center','margin-left':'0.2vw','fontSize':'1.2vw','color':'rgba(255,255,255,0.9)','height':'10%'}),
                                                    html.P(id='output_nbvage70',style={'textAlign': 'center','margin-left':'0.2vw','display':'block','fontWeight': 'bold','fontSize':'2vw','color':'rgba(255,255,255,0.9)','height':'20%'}),
                                                    ],className='card-body align-items-center justify-content-right',style={'margin-left':'11vw','margin-top':'2vw','display':'block',"border-radius":"100%","border":"1px rgba(255,255,255) solid",'fontWeight': 'bold','height':'90%','background-color':'rgba(0,0,0,0)','width':'35%'}),

                                                html.Div([
                                                    html.P(id='output_ucage70',style={'textAlign': 'center','margin-left':'0vw','margin-top':'-6vw','display':'block','fontWeight': 'bold','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'20%','width':'45%'}),
                                                    html.P("d'UC",style={'textAlign': 'center','margin-left':'-1.6vw','fontSize':'1vw','margin-top':'-5.4vw','color':'rgba(255,255,255,0.9)','height':'8%'}),
                                                    ],className='card-body d-flex justify-content-left',style={'display':'inline-block','height':'0%'}),

                                                html.Div([
                                                    html.P("Durée :",style={'textAlign': 'center','margin-left':'-0.2vw','fontSize':'1vw','margin-top':'-4.3vw','color':'rgba(255,255,255,0.9)','height':'8%'}),
                                                    html.P(id='output_durage70',style={'textAlign': 'center','margin-left':'-1vw','margin-top':'-4.8vw','display':'block','fontWeight': 'bold','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'20%','width':'45%'}),
                                                    ],className='card-body d-flex justify-content-left',style={'display':'inline-block','height':'16%'}),

                                                html.Div([
                                                    html.I(className="fas fa-chevron-right",style={'textAlign':'center','margin-top':'-7.5vw','margin-left':'8.5vw','fontSize':'2vw','color':'rgba(255,255,255,0.9)','height':'3vw'}),
                                                    ],className='card-body d-flex justify-content-left',style={'display':'inline-block','height':'13%'}),

                                             ],style={'background': 'linear-gradient(to left, rgba(219,112,147,0.0), rgba(20,27,34,0.6))','height':'6vw','margin-top':'-2vw'}),
                                        html.Div([
                                            html.I(className="fas fa-money-bill",style={'textAlign':'center','margin-top':'1.3vw','fontSize':'1.3vw','color':'rgba(255,255,255,0.9)','height':'3vw'}),
                                            html.P(id='output_patage70',style={'textAlign': 'center','margin-left':'-0.5vw','margin-top':'0.6vw','display':'block','fontWeight': 'bold','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'20%','width':'45%'}),
                                            html.P("du CA Confirmé",style={'textAlign': 'center','margin-left':'-1.8vw','fontSize':'1vw','margin-top':'1.10vw','color':'rgba(255,255,255,0.9)','height':'8%'}),
                                            ],className='card-body d-flex justify-content-left',style={'display':'inline-block','height':'13%'}),

                                        html.Div([
                                            html.I(className="fas fa-gas-pump",style={'textAlign':'center','margin-top':'1.8vw','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'3vw'}),
                                            html.P(id='output_nbage70',style={'textAlign': 'center','margin-left':'0.8vw','margin-top':'1.4vw','display':'block','fontWeight': 'bold','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'20%','width':'45%'}),
                                            html.P("contrats",style={'textAlign': 'center','margin-left':'-0.9vw','fontSize':'1vw','margin-top':'1.8vw','color':'rgba(255,255,255,0.9)','height':'8%'}),
                                            ],className='card-body d-flex justify-content-left',style={'display':'inline-block','height':'15%'}),
                                        
                                        html.Div([
                                            html.I(className="fas fa-credit-card mx-auto",style={'textAlign':'center','margin-top':'1.8vw','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'3vw'}),
                                            html.P(id='output_activiteage70',style={'textAlign': 'center','margin-left':'-0.3vw','margin-top':'1.5vw','display':'block','fontWeight': 'bold','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'20%','width':'45%'}),
                                            html.H3("alimenté", className = "card-body d-flex justify-content-center",style={'textAlign': 'center','margin-top':'1.4vw','margin-left':'-7vw','fontSize':'1vw','color':'rgba(255,255,255,0.9)'})
                                            ],className='card-body d-flex justify-content-center',style={'display':'inline-block','height':'15%'}),
                                        ])],color='rgba(57,62,69,0.9)',inverse=True,style={'border':'none','height':'23vw'} #
                                )
                            ])
                        ]),

                        dbc.Col([
                            html.Div([
                                dbc.Card([
                                    dbc.CardHeader(
                                       html.Div([
                                            html.I(className="fas fa-user",style={'margin-top':'0.5vw','margin-left':'0.5vw','textAlign':'center','fontSize':'2.5vw','color':'rgba(255,255,255,0.9)','height':'3vw','display':'inline-block'}),
                                            html.H3("Patrimoniaux - 75-80ans", className = "card-title mx-auto",style={'margin-top':'0.5vw','textAlign':'center','fontSize':'1.25vw','color':'rgba(255,255,255,0.9)','display':'inline-block'}),
                                            
                                            ],className='d-flex align-items-center justify-content-center'),
                                        style={'background': 'linear-gradient(to right, rgba(219,112,147,0.0), rgba(219,112,147,0.6))','height':'4.5vw'} 
                                        ),
                                    dbc.CardBody([
                                        html.Div([
                                                html.Div([
                                                    html.P("Valeur",style={'textAlign': 'center','margin-left':'0.2vw','fontSize':'1.2vw','color':'rgba(255,255,255,0.9)','height':'10%'}),
                                                    html.P(id='output_nbvage75',style={'textAlign': 'center','margin-left':'0.2vw','display':'block','fontWeight': 'bold','fontSize':'2vw','color':'rgba(255,255,255,0.9)','height':'20%'}),
                                                    ],className='card-body align-items-center justify-content-right',style={'margin-left':'11vw','margin-top':'2vw','display':'block',"border-radius":"100%","border":"1px rgba(255,255,255) solid",'fontWeight': 'bold','height':'90%','background-color':'rgba(0,0,0,0)','width':'35%'}),

                                                html.Div([
                                                    html.P(id='output_ucage75',style={'textAlign': 'center','margin-left':'0vw','margin-top':'-6vw','display':'block','fontWeight': 'bold','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'20%','width':'45%'}),
                                                    html.P("d'UC",style={'textAlign': 'center','margin-left':'-1.6vw','fontSize':'1vw','margin-top':'-5.4vw','color':'rgba(255,255,255,0.9)','height':'8%'}),
                                                    ],className='card-body d-flex justify-content-left',style={'display':'inline-block','height':'0%'}),

                                                html.Div([
                                                    html.P("Durée :",style={'textAlign': 'center','margin-left':'-0.2vw','fontSize':'1vw','margin-top':'-4.3vw','color':'rgba(255,255,255,0.9)','height':'8%'}),
                                                    html.P(id='output_durage75',style={'textAlign': 'center','margin-left':'-1vw','margin-top':'-4.8vw','display':'block','fontWeight': 'bold','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'20%','width':'45%'}),
                                                    ],className='card-body d-flex justify-content-left',style={'display':'inline-block','height':'16%'}),

                                                html.Div([
                                                    html.I(className="fas fa-chevron-right",style={'textAlign':'center','margin-top':'-7.5vw','margin-left':'8.5vw','fontSize':'2vw','color':'rgba(255,255,255,0.9)','height':'3vw'}),
                                                    ],className='card-body d-flex justify-content-left',style={'display':'inline-block','height':'13%'}),

                                             ],style={'background': 'linear-gradient(to left, rgba(219,112,147,0.0), rgba(20,27,34,0.6))','height':'6vw','margin-top':'-2vw'}),
                                        html.Div([
                                            html.I(className="fas fa-money-bill",style={'textAlign':'center','margin-top':'1.3vw','fontSize':'1.3vw','color':'rgba(255,255,255,0.9)','height':'3vw'}),
                                            html.P(id='output_patage75',style={'textAlign': 'center','margin-left':'-0.5vw','margin-top':'0.6vw','display':'block','fontWeight': 'bold','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'20%','width':'45%'}),
                                            html.P("du CA Confirmé",style={'textAlign': 'center','margin-left':'-1.8vw','fontSize':'1vw','margin-top':'1.10vw','color':'rgba(255,255,255,0.9)','height':'8%'}),
                                            ],className='card-body d-flex justify-content-left',style={'display':'inline-block','height':'13%'}),

                                        html.Div([
                                            html.I(className="fas fa-gas-pump",style={'textAlign':'center','margin-top':'1.8vw','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'3vw'}),
                                            html.P(id='output_nbage75',style={'textAlign': 'center','margin-left':'0.8vw','margin-top':'1.4vw','display':'block','fontWeight': 'bold','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'20%','width':'45%'}),
                                            html.P("contrats",style={'textAlign': 'center','margin-left':'-0.9vw','fontSize':'1vw','margin-top':'1.8vw','color':'rgba(255,255,255,0.9)','height':'8%'}),
                                            ],className='card-body d-flex justify-content-left',style={'display':'inline-block','height':'15%'}),
                                        
                                        html.Div([
                                            html.I(className="fas fa-credit-card mx-auto",style={'textAlign':'center','margin-top':'1.8vw','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'3vw'}),
                                            html.P(id='output_activiteage75',style={'textAlign': 'center','margin-left':'-0.3vw','margin-top':'1.5vw','display':'block','fontWeight': 'bold','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'20%','width':'45%'}),
                                            html.H3("alimenté", className = "card-body d-flex justify-content-center",style={'textAlign': 'center','margin-top':'1.4vw','margin-left':'-7vw','fontSize':'1vw','color':'rgba(255,255,255,0.9)'})
                                            ],className='card-body d-flex justify-content-center',style={'display':'inline-block','height':'15%'}),
                                        ])],color='rgba(69,68,86,0.9)',inverse=True,style={'border':'none','height':'23vw'} #
                                )
                            ])
                        ]),

                        dbc.Col([
                            html.Div([
                                dbc.Card([
                                    dbc.CardHeader(
                                        html.Div([
                                            html.I(className="fas fa-user",style={'margin-top':'0.5vw','margin-left':'0.5vw','textAlign':'center','fontSize':'2.5vw','color':'rgba(255,255,255,0.9)','height':'3vw','display':'inline-block'}),
                                            html.H3("Patrimoniaux -80ans", className = "card-title mx-auto",style={'margin-top':'0.5vw','textAlign':'center','fontSize':'1.25vw','color':'rgba(255,255,255,0.9)','display':'inline-block'}),
                                            
                                            ],className='d-flex align-items-center justify-content-center'),
                                        style={'background': 'linear-gradient(to right, rgba(219,112,147,0.0), rgba(219,112,147,0.6))','height':'4.5vw'} 
                                        ),
                                    dbc.CardBody([
                                       html.Div([
                                                html.Div([
                                                    html.P("Valeur",style={'textAlign': 'center','margin-left':'0.2vw','fontSize':'1.2vw','color':'rgba(255,255,255,0.9)','height':'10%'}),
                                                    html.P(id='output_nbvage80',style={'textAlign': 'center','margin-left':'0.2vw','display':'block','fontWeight': 'bold','fontSize':'2vw','color':'rgba(255,255,255,0.9)','height':'20%'}),
                                                    ],className='card-body align-items-center justify-content-right',style={'margin-left':'11vw','margin-top':'2vw','display':'block',"border-radius":"100%","border":"1px rgba(255,255,255) solid",'fontWeight': 'bold','height':'90%','background-color':'rgba(0,0,0,0)','width':'35%'}),

                                                html.Div([
                                                    html.P(id='output_ucage80',style={'textAlign': 'center','margin-left':'0vw','margin-top':'-6vw','display':'block','fontWeight': 'bold','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'20%','width':'45%'}),
                                                    html.P("d'UC",style={'textAlign': 'center','margin-left':'-1.6vw','fontSize':'1vw','margin-top':'-5.4vw','color':'rgba(255,255,255,0.9)','height':'8%'}),
                                                    ],className='card-body d-flex justify-content-left',style={'display':'inline-block','height':'0%'}),

                                                html.Div([
                                                    html.P("Durée :",style={'textAlign': 'center','margin-left':'-0.2vw','fontSize':'1vw','margin-top':'-4.3vw','color':'rgba(255,255,255,0.9)','height':'8%'}),
                                                    html.P(id='output_durage80',style={'textAlign': 'center','margin-left':'-1vw','margin-top':'-4.8vw','display':'block','fontWeight': 'bold','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'20%','width':'45%'}),
                                                    ],className='card-body d-flex justify-content-left',style={'display':'inline-block','height':'16%'}),

                                                html.Div([
                                                    html.I(className="fas fa-chevron-right",style={'textAlign':'center','margin-top':'-7.5vw','margin-left':'8.5vw','fontSize':'2vw','color':'rgba(255,255,255,0.9)','height':'3vw'}),
                                                    ],className='card-body d-flex justify-content-left',style={'display':'inline-block','height':'13%'}),

                                             ],style={'background': 'linear-gradient(to left, rgba(219,112,147,0.0), rgba(20,27,34,0.6))','height':'6vw','margin-top':'-2vw'}),
                                        html.Div([
                                            html.I(className="fas fa-money-bill",style={'textAlign':'center','margin-top':'1.3vw','fontSize':'1.3vw','color':'rgba(255,255,255,0.9)','height':'3vw'}),
                                            html.P(id='output_patage80',style={'textAlign': 'center','margin-left':'-0.5vw','margin-top':'0.6vw','display':'block','fontWeight': 'bold','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'20%','width':'45%'}),
                                            html.P("du CA Confirmé",style={'textAlign': 'center','margin-left':'-1.8vw','fontSize':'1vw','margin-top':'1.10vw','color':'rgba(255,255,255,0.9)','height':'8%'}),
                                            ],className='card-body d-flex justify-content-left',style={'display':'inline-block','height':'13%'}),

                                        html.Div([
                                            html.I(className="fas fa-gas-pump",style={'textAlign':'center','margin-top':'1.8vw','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'3vw'}),
                                            html.P(id='output_nbage80',style={'textAlign': 'center','margin-left':'0.8vw','margin-top':'1.4vw','display':'block','fontWeight': 'bold','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'20%','width':'45%'}),
                                            html.P("contrats",style={'textAlign': 'center','margin-left':'-0.9vw','fontSize':'1vw','margin-top':'1.8vw','color':'rgba(255,255,255,0.9)','height':'8%'}),
                                            ],className='card-body d-flex justify-content-left',style={'display':'inline-block','height':'15%'}),
                                        
                                        html.Div([
                                            html.I(className="fas fa-credit-card mx-auto",style={'textAlign':'center','margin-top':'1.8vw','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'3vw'}),
                                            html.P(id='output_activiteage80',style={'textAlign': 'center','margin-left':'-0.3vw','margin-top':'1.5vw','display':'block','fontWeight': 'bold','fontSize':'1.5vw','color':'rgba(255,255,255,0.9)','height':'20%','width':'45%'}),
                                            html.H3("alimenté", className = "card-body d-flex justify-content-center",style={'textAlign': 'center','margin-top':'1.4vw','margin-left':'-7vw','fontSize':'1vw','color':'rgba(255,255,255,0.9)'})
                                            ],className='card-body d-flex justify-content-center',style={'display':'inline-block','height':'15%'}),
                                        ])],color='rgba(67,93,94,0.9)',inverse=True,style={'border':'none','height':'23vw'} #
                                )
                            ])
                        ]),





                    ],className='d-flex align-items-stretch justify-content-center')
                    ],fluid=True)

            ], className='d-flex align-items-stretch justify-content-center'),
                        
            html.Br(),
            ])
        ])
    ]),className='d-flex align-items-center',color = 'rgba(44, 62, 80, 0.9)')
])


@app.callback(
    [Output("output_container", "children"),Output("output_uc", "children"),Output("output_age", "children"),Output("output_turnover", "children"),
    Output("output_activ", "children"),Output("output_prime", "children"),
    Output("pie_chart", "figure"),Output("scatter_chart", "figure"),Output("nbpc", "children"),Output("nbgp", "children"),Output("nbpp", "children"),
    Output("gb0", "value"),Output("nbglo", "children"),Output("nbeu", "children"),Output("nbuc", "children")], 
    [Input("radio_value", "value")])

def update_pie(mark_selected):
    dfmp_filtered = dfmp.loc[dfmp['Marque'] == mark_selected].sort_values(by=['Pat'])
    dfmaps_filtered = dfmaps.loc[dfmaps['Marque'] == mark_selected]
    dfcarac_filtered = dfcarac.loc[dfcarac['Marque'] == mark_selected]
    dfmpmk_filtered = dfmpmk.loc[dfmpmk['Marque'] == mark_selected]
    df=dfcoll.loc[dfcoll['Marque'] == mark_selected]
    dfuc=df.loc[df['Segment']=="UC"]
    
    card1 = int(np.round(df["collecte"].sum()/10**6,0))
    card2 = int(np.round(dfuc["collecte"].sum()/df["collecte"].sum()*100,0))
    card3 = dfmp_filtered["Age_marque"].max()
    card4 = dfmp_filtered["Prime_marque"].max()
    card5 = dfcarac_filtered["Activ"].max()
    card6 = int(np.round(df["prestations"].sum()/df["PM actu"].sum()*100,1))

    valeur_UC = dfmpmk_filtered['%UC']
    valeur_PC = dfmp_filtered.loc[dfmp_filtered["Pat"] == 'Confirme']["pct_nbv"].max()
    valeur_GP = dfmp_filtered.loc[dfmp_filtered["Pat"] == 'Grand Public']["pct_nbv"].max()
    valeur_PP = dfmp_filtered.loc[dfmp_filtered["Pat"] == 'Potentiel']["pct_nbv"].max()
    valeur_glo_tot = dfmpmk_filtered['NBV'].max()
    valeur_glo_eu = dfmpmk_filtered['NBV_eu'].max()
    valeur_glo_uc = dfmpmk_filtered['NBV_uc'].max()
       
 
    nbv_conf = "PC : {}%".format(valeur_PC)
    nbv_gp = "GP : {}%".format(valeur_GP)
    nbv_pp = "PP : {}%".format(valeur_PP)
    nbv_glo = "{}%".format(valeur_glo_tot)
    nbv_eu = "€ : {}%".format(valeur_glo_eu)
    nbv_uc = "UC : {}%".format(valeur_glo_uc)

    gb_uc = valeur_UC*100



    container = "{} M€".format(card1)
    uc = "{}%".format(card2)
    age = "{} ans".format(card3)
    prime = "{} K€".format(card4)
    activ = "1 sur {}".format(card5)
    turnover = "{} %".format(card6)

    fig=px.pie(dfmp_filtered, values="%CA", names="Pat", hole=0.6,color_discrete_sequence=color4,opacity=1
                    ).update_layout(
                        template='plotly_dark',
                        plot_bgcolor= 'rgba(0, 0, 0, 0)',
                        paper_bgcolor= 'rgba(0, 0, 0, 0)',
                           
                        uniformtext_minsize=10, 
                        uniformtext_mode='hide',
                        #height=400,
                        margin=dict(l=0,r=60,b=80,t=0),
                        #showlegend=False,
                        legend=dict(
                            title='',
                            tracegroupgap = 0.5,
                            orientation='v',
                            font=dict(size=10),
                            yanchor='auto',
                            xanchor='auto',
                            x=-0.4,
                            y=-0.0)
                    ).update_traces(
                      textposition = 'inside',     
                    )                       
                    


    fig3=px.line(dfmaps_filtered, x="Trage", y="%CA_PAT", color="PAT",
                        color_discrete_sequence=color4,
                        line_shape='spline',
                    ).update_traces(#fill='tonexty',
                        line_width=4,
                        mode='lines+markers',
                        marker=dict(size=9,opacity=1)
                    ).update_layout(
                        template='plotly_dark',
                        plot_bgcolor= 'rgba(0, 0, 0, 0)',
                        paper_bgcolor= 'rgba(0, 0, 0, 0)',
                        #height=350,
                        showlegend=False,
                        xaxis_title="",
                        yaxis_title="",
                        yaxis_tickformat = '.0%',
                    )
    
    fig4=px.line(dfmaps_filtered, x='Trage', y='%NBV', color='PAT',
                    color_discrete_sequence=color4,line_shape='spline',
                    ).update_layout(
                        title_x=0.48,title_y=0.96,
                        yaxis_title="NBV (%CA)",
                        xaxis_title="",
                        yaxis_tickformat = '.1%',
                        template='plotly_dark',
                        plot_bgcolor= 'rgba(0, 0, 0, 0)',
                        paper_bgcolor= 'rgba(0, 0, 0, 0)',
                        #height=350,
                        xaxis = dict(tickfont = dict(size=11)),
                        showlegend=False,
                        margin=dict(l=10,r=0,b=130,t=0),
                        ).update_traces(
                            line_width=4,
                            mode='lines+markers',
                            marker=dict(size=12,opacity=0.8)
                        )
                        


    return container,uc,age,turnover,activ,prime,fig,fig4,nbv_conf,nbv_gp,nbv_pp,gb_uc,nbv_glo,nbv_eu,nbv_uc

@app.callback(
Output("output_patage40", "children"),Output("output_ucage40", "children"),Output("output_nbvage40", "children"),Output("output_nbage40", "children"),Output("output_activiteage40", "children"),Output("output_durage40", "children"),
Output("output_patage70", "children"),Output("output_ucage70", "children"),Output("output_nbvage70", "children"),Output("output_nbage70", "children"),Output("output_activiteage70", "children"),Output("output_durage70", "children"),
Output("output_patage75", "children"),Output("output_ucage75", "children"),Output("output_nbvage75", "children"),Output("output_nbage75", "children"),Output("output_activiteage75", "children"),Output("output_durage75", "children"),
Output("output_patage80", "children"),Output("output_ucage80", "children"),Output("output_nbvage80", "children"),Output("output_nbage80", "children"),Output("output_activiteage80", "children"),Output("output_durage80", "children"),
[Input("radio_value", "value")])


def update_nbvage(mark_selected):
    df = dfcompl[dfcompl['Marque']==mark_selected]
    dfpat=df[df['SEGMENT_MODOP']=="PC"]
    dfuc=dfcoll[dfcoll['Marque']==mark_selected]
    dfuc=dfuc[dfuc['SEGMENT_MODOP']=="PC"]
    dfuc=dfuc[dfuc['Trage']=="40-50ans"]   
    dfuc=dfuc[dfuc['Segment']=="UC"]
    df = df[df['reporting+70']==1]
    df = df[df['SEGMENT_MODOP']=="PC"]
    df['NBV']=df['PVFP']/df['collecte']
    df=df[df['Trage']=="40-50ans"]
    nbvage40=int(np.round((df["NBV"].sum())*100,1))
    nbvage40 = "{}%".format(nbvage40)
    collage40=int(np.round((df["collecte"].sum()/1000000),0))
    collage40 = "{}M€".format(collage40)
    patage40=int(np.round((df["collecte"].sum()/dfpat["collecte"].sum())*100,1))
    patage40 = "{}%".format(patage40)
    ucage40=int(np.round((dfuc["collecte"].sum()/df["collecte"].sum())*100,0))
    ucage40 = "{}%".format(ucage40)
    nbage40=df['NBSTIONS'].sum()
    nbage40 = "{}".format(nbage40)
    activite40=int(np.round((df["NBSTIONS"].sum()/df["NBVERSEURS"].sum()),0))
    activite40 = "1/{}".format(activite40)
    durage40=int(np.round(df['PM actu'].sum()/df['collecte'].sum(),0))
    durage40 = "{} ans".format(durage40)
  

    df = dfcompl[dfcompl['Marque']==mark_selected]
    dfpat=df[df['SEGMENT_MODOP']=="PC"]
    dfuc=dfcoll[dfcoll['Marque']==mark_selected]
    dfuc=dfuc[dfuc['SEGMENT_MODOP']=="PC"]
    dfuc=dfuc[dfuc['Trage']=="70-75ans"]  
    dfuc=dfuc[dfuc['reporting+70']==1]
    dfuc=dfuc[dfuc['Segment']=="UC"]
    df = df[df['reporting+70']==1]
    df = df[df['SEGMENT_MODOP']=="PC"]
    df['NBV']=df['PVFP']/df['collecte']
    df=df[df['Trage']=="70-75ans"]
    nbvage70=int(np.round((df["NBV"].sum())*100,2))
    nbvage70 = "{}%".format(nbvage70)
    collage70=int(np.round((df["collecte"].sum()/1000000),0))
    collage70 = "{}M€".format(collage70)
    patage70=int(np.round((df["collecte"].sum()/dfpat["collecte"].sum())*100,1))
    patage70 = "{}%".format(patage70)
    ucage70=int(np.round((dfuc["collecte"].sum()/df["collecte"].sum())*100,0))
    ucage70 = "{}%".format(ucage70)
    nbage70=df['NBSTIONS'].sum()
    nbage70 = "{}".format(nbage70)
    activite70=int(np.round((df["NBSTIONS"].sum()/df["NBVERSEURS"].sum()),0))
    activite70 = "1/{}".format(activite70)
    durage70=int(np.round(df['PM actu'].sum()/df['collecte'].sum(),0))
    durage70 = "{} ans".format(durage70)


    df = dfcompl[dfcompl['Marque']==mark_selected]
    dfpat=df[df['SEGMENT_MODOP']=="PC"]
    dfuc=dfcoll[dfcoll['Marque']==mark_selected]
    dfuc=dfuc[dfuc['SEGMENT_MODOP']=="PC"]
    dfuc=dfuc[dfuc['Trage']=="75-80ans"]  
    dfuc=dfuc[dfuc['reporting+70']==1]
    dfuc=dfuc[dfuc['Segment']=="UC"]
    df = df[df['reporting+70']==1]
    df = df[df['SEGMENT_MODOP']=="PC"]
    df['NBV']=df['PVFP']/df['collecte']
    df=df[df['Trage']=="75-80ans"]
    nbvage75=int(np.round((df["NBV"].sum())*100,2))
    nbvage75 = "{}%".format(nbvage75)
    collage75=int(np.round((df["collecte"].sum()/1000000),0))
    collage75 = "{}M€".format(collage75)
    patage75=int(np.round((df["collecte"].sum()/dfpat["collecte"].sum())*100,1))
    patage75 = "{}%".format(patage75)
    ucage75=int(np.round((dfuc["collecte"].sum()/df["collecte"].sum())*100,0))
    ucage75 = "{}%".format(ucage75)
    nbage75=df['NBSTIONS'].sum()
    nbage75 = "{}".format(nbage75)
    activite75=int(np.round((df["NBSTIONS"].sum()/df["NBVERSEURS"].sum()),0))
    activite75 = "1/{}".format(activite75)
    durage75=int(np.round(df['PM actu'].sum()/df['collecte'].sum(),0))
    durage75 = "{} ans".format(durage75)


    df = dfcompl[dfcompl['Marque']==mark_selected]
    dfpat=df[df['SEGMENT_MODOP']=="PC"]
    dfuc=dfcoll[dfcoll['Marque']==mark_selected]
    dfuc=dfuc[dfuc['SEGMENT_MODOP']=="PC"]
    dfuc=dfuc[dfuc['Trage']=="+80ans"]  
    dfuc=dfuc[dfuc['reporting+70']==1]
    dfuc=dfuc[dfuc['Segment']=="UC"]
    df = df[df['reporting+70']==1]
    df = df[df['SEGMENT_MODOP']=="PC"]
    df['NBV']=df['PVFP']/df['collecte']
    df=df[df['Trage']=="+80ans"]
    nbvage80=int(np.round((df["NBV"].sum())*100,2))
    nbvage80 = "{}%".format(nbvage80)
    collage80=int(np.round((df["collecte"].sum()/1000000),0))
    collage80 = "{}M€".format(collage80)
    patage80=int(np.round((df["collecte"].sum()/dfpat["collecte"].sum())*100,1))
    patage80 = "{}%".format(patage80)
    ucage80=int(np.round((dfuc["collecte"].sum()/df["collecte"].sum())*100,0))
    ucage80 = "{}%".format(ucage80)
    nbage80=df['NBSTIONS'].sum()
    nbage80 = "{}".format(nbage80)
    activite80=int(np.round((df["NBSTIONS"].sum()/df["NBVERSEURS"].sum()),0))
    activite80 = "1/{}".format(activite80)
    durage80=int(np.round(df['PM actu'].sum()/df['collecte'].sum(),0))
    durage80 = "{} ans".format(durage80)





    return patage40,ucage40,nbvage40,nbage40,activite40,durage40,patage70,ucage70,nbvage70,nbage70,activite70,durage70,patage75,ucage75,nbvage75,nbage75,activite75,durage75,patage80,ucage80,nbvage80,nbage80,activite80,durage80


# Run app and display result inline in the notebook
if __name__ == "__main__":
    server=pywsgi.WSGIServer(('127.0.0.1',5000),app)
    server.serve_forever()
